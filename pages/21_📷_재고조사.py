# -*- coding: utf-8 -*-
"""
재고자산 실사 (사진/QR코드 대조)

- 엑셀 자산번호 목록을 불러오거나(엑셀 업로드), 자산관리 앱의 목록을 불러옵니다.
- 자산번호마다 QR코드 라벨 시트를 만들어서 인쇄 → 자산에 붙일 수 있습니다.
- 아이패드/휴대폰 카메라로 QR코드를 촬영하면, 목록에 있는 자산인지 바로 대조해서 보여줍니다.
- 확인된 자산은 계속 누적되어 진행 현황(확인 N / 전체 M)을 보여주고, 결과를 엑셀로 다운로드할 수 있습니다.
"""

import streamlit as st
import pandas as pd
import json
import os
import sys
import hashlib
from io import BytesIO
from datetime import datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from style import apply_common_style, show_page_header, check_password

st.set_page_config(page_title="재고조사 - 마이 유틸리티", page_icon="📷", layout="wide")
apply_common_style()
check_password()
show_page_header("📷", "재고자산 실사", "자산번호 목록을 불러오고, 사진(QR코드)을 찍어서 바로 대조하세요")

# ── 필요한 패키지 확인 ──────────────────────────────────────
try:
    import qrcode
    from PIL import Image, ImageDraw, ImageFont
    from pyzbar.pyzbar import decode as decode_barcode
    LIBS_OK = True
except ImportError:
    LIBS_OK = False

if not LIBS_OK:
    st.error(
        "이 기능을 쓰려면 필요한 프로그램(패키지)이 아직 설치되어 있지 않아요.\n\n"
        "터미널(명령 프롬프트)을 열어서 아래 명령을 실행한 뒤, 이 페이지를 새로고침 해주세요."
    )
    st.code("pip install qrcode pyzbar pillow", language="bash")
    st.stop()

ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
ASSET_JSON_PATH = os.path.join(ROOT_DIR, "asset_data.json")

# ── 세션 상태 초기화 ────────────────────────────────────────
st.session_state.setdefault("inv_checked", set())      # 확인된 자산번호 집합
st.session_state.setdefault("inv_unknown", [])          # 목록에 없는데 스캔된 항목
st.session_state.setdefault("inv_last_hash", None)      # 마지막으로 처리한 사진(중복 처리 방지)


# ── 자산관리 앱 데이터 읽기 (읽기 전용, asset_manager.py는 건드리지 않음) ──
def load_asset_manager_list():
    if not os.path.exists(ASSET_JSON_PATH):
        return None
    try:
        with open(ASSET_JSON_PATH, encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return None

    assets = data.get("assets", [])
    if not assets:
        return None

    rows = [{
        "자산ID": a.get("id", ""),
        "자산명": a.get("name", ""),
        "분류": a.get("category", ""),
        "위치": a.get("location", ""),
        "부서": a.get("department", ""),
        "사용자": a.get("current_user", ""),
        "상태": a.get("status", ""),
    } for a in assets]
    return pd.DataFrame(rows)


# ── 한글이 표시되는 폰트 찾기 (없으면 기본 폰트로 대체) ──────
def load_korean_font(size):
    for path in [
        "C:/Windows/Fonts/malgun.ttf",
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/System/Library/Fonts/AppleSDGothicNeo.ttc",
    ]:
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            continue
    return ImageFont.load_default()


# ── QR 라벨 시트 이미지 만들기 ──────────────────────────────
def make_label_sheet(items):
    """items: [(자산번호, 자산명), ...] → 인쇄용 라벨 시트 이미지(PIL Image) 반환"""
    cols = 3
    cell_w, cell_h = 260, 330
    margin = 20
    rows = -(-len(items) // cols)  # 올림 나눗셈

    sheet = Image.new("RGB", (cols * cell_w + margin * 2, rows * cell_h + margin * 2), "white")
    draw = ImageDraw.Draw(sheet)
    font_id = load_korean_font(18)
    font_name = load_korean_font(14)

    for i, (asset_id, asset_name) in enumerate(items):
        col, row = i % cols, i // cols
        x, y = margin + col * cell_w, margin + row * cell_h

        qr_img = qrcode.make(str(asset_id), box_size=5, border=2).convert("RGB")
        qr_size = 200
        qr_img = qr_img.resize((qr_size, qr_size))
        sheet.paste(qr_img, (x + (cell_w - qr_size) // 2, y + 10))

        text_y = y + 10 + qr_size + 8
        id_text = str(asset_id)
        tw = draw.textlength(id_text, font=font_id)
        draw.text((x + (cell_w - tw) / 2, text_y), id_text, fill="black", font=font_id)

        if asset_name:
            name_text = str(asset_name)
            tw2 = draw.textlength(name_text, font=font_name)
            draw.text((x + (cell_w - tw2) / 2, text_y + 24), name_text, fill="#636E72", font=font_name)

        draw.rectangle([x + 2, y + 2, x + cell_w - 2, y + cell_h - 2], outline="#DDDDDD")

    return sheet


# ── 실사 결과 엑셀 만들기 ────────────────────────────────────
def build_result_excel(df, id_col, checked_ids, unknown_list):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "실사결과"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    miss_fill = PatternFill(start_color="FDE2E2", end_color="FDE2E2", fill_type="solid")

    headers = list(df.columns) + ["확인여부"]
    for c, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=text)
        cell.font = header_font
        cell.fill = header_fill

    id_series = df[id_col].astype(str).str.strip()
    for r, (_, row) in enumerate(df.iterrows(), 2):
        for c, col_name in enumerate(df.columns, 1):
            ws.cell(row=r, column=c, value=row[col_name])
        checked = id_series.iloc[r - 2] in checked_ids
        cell = ws.cell(row=r, column=len(df.columns) + 1, value="확인됨" if checked else "미확인")
        if not checked:
            cell.fill = miss_fill

    if unknown_list:
        ws2 = wb.create_sheet("목록에없는스캔항목")
        ws2.append(["자산번호", "촬영시각"])
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
        for item in unknown_list:
            ws2.append([item["자산번호"], item["촬영시각"]])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── ① 대조 목록 준비 ─────────────────────────────────────────
st.markdown("### 1️⃣ 대조할 자산번호 목록 준비")

source = st.radio(
    "목록을 어디서 가져올까요?",
    ["엑셀 파일 업로드", "자산관리 앱 목록 사용"],
    horizontal=True,
)

df = None
if source == "엑셀 파일 업로드":
    uploaded = st.file_uploader("자산번호가 들어있는 엑셀 파일(.xlsx)을 올려주세요", type=["xlsx"])
    if uploaded is not None:
        try:
            df = pd.read_excel(uploaded, dtype=str)
        except Exception as e:
            st.error(f"엑셀 파일을 읽는 중 문제가 발생했어요: {e}")
else:
    df = load_asset_manager_list()
    if df is None:
        st.info("자산관리 앱(asset_manager.py)에 등록된 자산이 아직 없어요. '엑셀 파일 업로드'를 이용해주세요.")

if df is not None and not df.empty:
    cols = list(df.columns)

    def guess_col(keywords, fallback_idx=0):
        for c in cols:
            if any(k in str(c) for k in keywords):
                return c
        return cols[fallback_idx]

    id_default = guess_col(["자산번호", "자산ID", "ID", "번호"])
    name_default = guess_col(["자산명", "이름", "품명"], fallback_idx=min(1, len(cols) - 1))

    col1, col2 = st.columns(2)
    with col1:
        id_col = st.selectbox("어느 열이 '자산번호'인가요?", cols, index=cols.index(id_default))
    with col2:
        name_options = ["(사용 안 함)"] + cols
        name_col = st.selectbox(
            "라벨에 같이 표시할 '자산명' 열 (선택)", name_options,
            index=name_options.index(name_default) if name_default in name_options else 0,
        )
        name_col = None if name_col == "(사용 안 함)" else name_col

    id_series = df[id_col].astype(str).str.strip()
    st.success(f"목록을 불러왔어요. 총 {len(df)}개 자산")
    st.dataframe(df, use_container_width=True, height=200)

    st.divider()

    # ── ② QR 라벨 만들기 ─────────────────────────────────────
    st.markdown("### 2️⃣ QR 라벨 만들기 (자산에 붙일 라벨이 없다면 먼저 인쇄하세요)")
    st.caption("자산마다 QR코드가 인쇄된 라벨 시트를 만들어드려요. 인쇄해서 오린 뒤, 각 자산에 붙여주세요.")

    if st.button("🏷️ QR 라벨 시트 만들기"):
        names = df[name_col].astype(str).tolist() if name_col else [""] * len(id_series)
        items = list(zip(id_series.tolist(), names))
        with st.spinner("라벨 시트를 만들고 있어요..."):
            sheet = make_label_sheet(items)
        buf = BytesIO()
        sheet.save(buf, format="PNG")
        st.image(sheet, caption="라벨 시트 미리보기 (인쇄해서 오려 자산에 붙여주세요)")
        st.download_button(
            "📥 라벨 이미지 다운로드", data=buf.getvalue(),
            file_name="자산_QR라벨.png", mime="image/png",
        )

    st.divider()

    # ── ③ 사진 촬영 → 대조 ───────────────────────────────────
    st.markdown("### 3️⃣ 사진 촬영해서 대조하기")
    st.caption("라벨을 붙인 자산의 QR코드를 촬영하면, 목록에 있는지 바로 확인해드려요.")

    photo = st.camera_input("자산의 QR코드를 촬영하세요 (가까이서, 밝은 곳에서 찍어주세요)")

    if photo is not None:
        photo_hash = hashlib.md5(photo.getvalue()).hexdigest()
        if photo_hash != st.session_state.inv_last_hash:
            st.session_state.inv_last_hash = photo_hash

            img = Image.open(photo)
            img.thumbnail((1200, 1200))
            results = decode_barcode(img)

            if not results:
                st.warning("바코드(QR코드)를 찾지 못했어요. 더 가까이서, 밝은 곳에서 다시 찍어주세요.")
            else:
                code = results[0].data.decode("utf-8", errors="ignore").strip()
                match_rows = df[id_series == code]
                if not match_rows.empty:
                    st.session_state.inv_checked.add(code)
                    row = match_rows.iloc[0]
                    st.success(f"✅ 확인됨: {code}")
                    st.dataframe(match_rows, use_container_width=True)
                else:
                    st.session_state.inv_unknown.append({
                        "자산번호": code,
                        "촬영시각": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    })
                    st.error(f"❌ 목록에 없는 자산번호입니다: {code}")

    st.divider()

    # ── ④ 진행 현황 + 결과 다운로드 ──────────────────────────
    st.markdown("### 4️⃣ 진행 현황")

    total = len(df)
    checked_count = sum(1 for v in id_series if v in st.session_state.inv_checked)
    st.progress(checked_count / total if total else 0)
    st.write(f"✅ 확인 {checked_count} / 전체 {total}건")

    unchecked_df = df[~id_series.isin(st.session_state.inv_checked)]
    with st.expander(f"아직 확인 안 된 자산 ({len(unchecked_df)}건)"):
        st.dataframe(unchecked_df, use_container_width=True)

    if st.session_state.inv_unknown:
        with st.expander(f"목록에 없는데 스캔된 자산번호 ({len(st.session_state.inv_unknown)}건)"):
            st.dataframe(pd.DataFrame(st.session_state.inv_unknown), use_container_width=True)

    col_a, col_b = st.columns(2)
    with col_a:
        excel_bytes = build_result_excel(df, id_col, st.session_state.inv_checked, st.session_state.inv_unknown)
        st.download_button(
            "📥 실사 결과 엑셀로 저장", data=excel_bytes,
            file_name=f"재고조사결과_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with col_b:
        if st.button("🔄 처음부터 다시"):
            st.session_state.inv_checked = set()
            st.session_state.inv_unknown = []
            st.session_state.inv_last_hash = None
            st.rerun()

elif df is None:
    st.info("먼저 자산번호 목록을 불러와주세요.")
else:
    st.warning("목록에 데이터가 없어요. 엑셀 파일 내용을 확인해주세요.")
